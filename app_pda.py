import os
import time
import json
import threading
from collections import deque
from typing import Dict, List, Tuple

import requests
from flask import Flask, request, jsonify, send_from_directory

try:
    import openpyxl
except Exception:
    openpyxl = None

APP_VERSION = "2.1.2"
BASE = "https://dispatch-api.uniuni.com"

URL_LOGIN = f"{BASE}/map/login"
URL_ORDER_DETAIL = f"{BASE}/map/getorderdetail"
URL_ASSIGN_DRIVER = f"{BASE}/map/assignorderstodriver"
URL_UPDATE_STATUS = f"{BASE}/driver/updateshippingstatus"
URL_OPERATION_LOG = f"{BASE}/driver/insertoperationlog"
URL_BATCH_LIST = f"{BASE}/map/getdispatchlisthistory"
URL_QUICK_TRANSFER = f"{BASE}/business/quicktransferorders"

WAREHOUSE_NAME_MAP = {
    1: "LAX Warehouse",
    31: "ATL Warehouse",
    44: "SAV Warehouse",
    46: "CHS Warehouse",
    50: "BNA Warehouse",
    60: "TYS Warehouse",
    61: "GSP Warehouse",
    62: "CAE Warehouse",
    66: "BFM Warehouse",
    67: "BHM Warehouse",
    76: "JAN Warehouse",
}

SKIP_STATUS_UPDATE = {195, 199, 200, 202, 231, 232}
STORED_DRIVER = 310991   # fixed driver for 213/230 stored parcels
CURRENT_USER = ""


def get_998_driver(warehouse: int) -> int:
    if warehouse == 31:
        return 310998
    return int(str(warehouse) + "0998")


class TokenManager:
    def __init__(self):
        self._lock = threading.Lock()
        self._username = ""
        self._password = ""
        self._token = ""

    def set_credentials(self, username: str, password: str):
        with self._lock:
            self._username = (username or "").strip()
            self._password = password or ""
            self._token = ""

    def is_ready(self) -> bool:
        with self._lock:
            return bool(self._username and self._password)

    def _login(self) -> str:
        r = requests.post(
            URL_LOGIN,
            params={"username": self._username},
            json={"password": self._password},
            timeout=15,
        )
        r.raise_for_status()
        j = r.json()
        if j.get("status") != "SUCCESS" or "token" not in j.get("data", {}):
            raise RuntimeError(f"Login failed: {j.get('ret_msg') or j.get('err_code')}")

        global CURRENT_USER
        data = j["data"]
        CURRENT_USER = (data.get("username") or self._username or "").strip()
        self._token = data["token"]
        return self._token

    def force_refresh(self) -> str:
        with self._lock:
            self._token = ""
            return self._login()

    def get_token(self) -> str:
        with self._lock:
            if not self._token:
                return self._login()
            return self._token


token_mgr = TokenManager()


def api_req(method, url, *, params=None, body=None, timeout=25):
    token = token_mgr.get_token()
    headers = {
        "Authorization": "Bearer " + token,
        "Content-Type": "application/json",
    }
    r = requests.request(method, url, params=params, json=body, headers=headers, timeout=timeout)
    if r.status_code in (401, 403):
        token = token_mgr.force_refresh()
        headers["Authorization"] = "Bearer " + token
        r = requests.request(method, url, params=params, json=body, headers=headers, timeout=timeout)
    return r


def must_json(r, name):
    if r.status_code != 200:
        raise RuntimeError(f"{name} HTTP {r.status_code}: {r.text[:300]}")
    try:
        return r.json()
    except Exception:
        raise RuntimeError(f"{name} non-JSON")


def _as_dict(obj):
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, str):
        try:
            v = json.loads(obj)
            return v if isinstance(v, dict) else {}
        except Exception:
            return {}
    return {}


def _data(j):
    root = _as_dict(j)
    d = root.get("data")
    if isinstance(d, str):
        d = _as_dict(d)
    return d if isinstance(d, dict) else {}


ROUTE_XLSX = os.environ.get("ROUTE_XLSX", "route.xlsx")
_route_lock = threading.Lock()
_zip_to_route: Dict[str, int] = {}
_route_loaded = False
_route_msg = "not loaded"


def _norm_zip(z):
    digits = "".join(c for c in (z or "") if c.isdigit())
    return digits[:5] if len(digits) >= 5 else digits


def load_routes():
    global _zip_to_route, _route_loaded, _route_msg

    if openpyxl is None:
        with _route_lock:
            _route_loaded = False
            _route_msg = "openpyxl not installed"
        return False, _route_msg, 0

    path = os.path.abspath(ROUTE_XLSX)
    if not os.path.exists(path):
        with _route_lock:
            _route_loaded = False
            _route_msg = f"file not found: {path}"
        return False, _route_msg, 0

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        z2r = {}

        for sheet in wb.worksheets:
            headers = {}
            for c in range(1, sheet.max_column + 1):
                v = sheet.cell(row=1, column=c).value
                if v is not None:
                    headers[str(v).strip().lower()] = c

            if "zipcode" not in headers or "route_no" not in headers:
                continue

            col_zip = headers["zipcode"]
            col_route = headers["route_no"]
            col_en = headers.get("is_enabled")

            for row in range(2, sheet.max_row + 1):
                rz = sheet.cell(row=row, column=col_zip).value
                rr = sheet.cell(row=row, column=col_route).value
                if rz is None or rr is None:
                    continue

                if col_en is not None:
                    try:
                        if int(sheet.cell(row=row, column=col_en).value) != 1:
                            continue
                    except Exception:
                        pass

                z = _norm_zip(str(rz))
                if not z:
                    continue

                try:
                    z2r[z] = int(str(rr).strip())
                except Exception:
                    pass

        with _route_lock:
            _zip_to_route = z2r
            _route_loaded = True
            _route_msg = f"loaded {len(z2r)} zip codes"

        return True, _route_msg, len(z2r)

    except Exception as e:
        with _route_lock:
            _route_loaded = False
            _route_msg = f"load error: {e}"
        return False, _route_msg, 0


def route_lookup(zip5):
    with _route_lock:
        return _zip_to_route.get(_norm_zip(zip5))


def get_blind_batch(warehouse):
    try:
        r = api_req(
            "GET",
            URL_BATCH_LIST,
            params={"branch": warehouse, "page": 1, "offset": 200},
            timeout=20,
        )
        j = must_json(r, "getdispatchlisthistory")
        items = (j.get("data") or {}).get("data") or []

        candidates = [
            x for x in items
            if x.get("is_removed") == 0 and "blind" in (x.get("name") or "").lower()
        ]
        if not candidates:
            return None, None

        best = max(candidates, key=lambda x: x.get("create_time", 0))

        batch_no = (
            best.get("batch_no")
            or best.get("sub_batch_no")
            or best.get("dispatch_no")
            or None
        )

        if not batch_no:
            details = (best.get("dispatch_details") or "").strip()
            batch_no = details.split(",")[-1].strip() if details else None

        return batch_no, best.get("name") or ""
    except Exception:
        return None, None


def get_order_detail(tno):
    r = api_req("GET", URL_ORDER_DETAIL, params={"tno": tno}, timeout=20)
    return must_json(r, "getorderdetail")


def parse_order(detail):
    d = _data(detail)
    orders = _as_dict(d.get("orders") or {})
    tracking = _as_dict(d.get("tracking") or {})

    order_id = warehouse = driver_id = state = frt = None
    try:
        order_id = int(orders.get("order_id"))
    except Exception:
        pass
    try:
        warehouse = int(orders.get("warehouse"))
    except Exception:
        pass
    try:
        driver_id = int(orders.get("shipping_staff_id") or 0) or None
    except Exception:
        pass
    try:
        state = int(tracking.get("state") or orders.get("latest_status"))
    except Exception:
        pass
    try:
        frt = int(orders.get("failed_reason_type") or -1)
    except Exception:
        pass

    storage_info = (
        str(tracking.get("storage_info") or orders.get("storage_info") or "").strip()
    )

    return {
        "order_id": order_id,
        "warehouse": warehouse,
        "driver_id": driver_id,
        "state": state,
        "failed_reason_type": frt,
        "zip": str(orders.get("zipcode") or tracking.get("zip") or "").strip(),
        "address": str(tracking.get("address1") or orders.get("address") or "").strip(),
        "city": str(tracking.get("city") or "").strip(),
        "province": str(tracking.get("province") or tracking.get("province_code") or "").strip(),
        "consignee": str(orders.get("consignee") or "").strip(),
        "tno": str(orders.get("tno") or "").strip(),
        "storage_info": storage_info,
    }


def do_assign_driver(order_id, from_driver, to_driver):
    payload = {
        "orders": [order_id],
        "to_driver": str(to_driver),
        "from_driver": int(from_driver or 0),
        "operator": CURRENT_USER or "unknown",
    }
    try:
        r = api_req("POST", URL_ASSIGN_DRIVER, body=payload, timeout=20)
        j = must_json(r, "assignorderstodriver")
        if j.get("status") == "SUCCESS":
            return True, "OK"
        return False, j.get("ret_msg") or str(j)
    except Exception as e:
        return False, str(e)


def do_insert_log(order_id, step):
    try:
        api_req(
            "POST",
            URL_OPERATION_LOG,
            body={
                "order_id": order_id,
                "operator": CURRENT_USER or "unknown",
                "operation_code": step,
                "operation_type": 0,
                "description": "",
                "memo": "",
            },
            timeout=15,
        )
    except Exception:
        pass
    time.sleep(0.1)


def do_update_202(order_id, warehouse, tno):
    scan_location = WAREHOUSE_NAME_MAP.get(warehouse, f"Warehouse {warehouse}")
    payload = {
        "order_id": order_id,
        "staff_id": 666,
        "operator": CURRENT_USER or "unknown",
        "shipping_status": 1,
        "scan_location": scan_location,
        "send_sms": 0,
        "storaged_warehouse": warehouse,
        "warehouse": warehouse,
        "warehouse_id": warehouse,
        "exception": None,
        "failed_reason": None,
        "parcel_info": {
            "order_id": order_id,
            "extra_order_sn": tno,
            "transition": "RE_TRANSIT",
            "status": 202,
        },
    }
    try:
        r = api_req("POST", URL_UPDATE_STATUS, body=payload, timeout=25)
        j = must_json(r, "updateshippingstatus-202")
        if j.get("status") == "SUCCESS":
            return True, "OK"
        return False, j.get("ret_msg") or str(j)
    except Exception as e:
        return False, str(e)


def do_transfer_to_batch(tno, batch_no, warehouse):
    payload = {
        "tnos": tno,
        "driver_id": "",
        "assign_to_sub_batch": True,
        "sub_batch_no": batch_no,
        "operator": CURRENT_USER or "unknown",
        "warehouse": warehouse,
    }
    try:
        r = api_req("POST", URL_QUICK_TRANSFER, body=payload, timeout=25)
        j = must_json(r, "quicktransferorders")
        if j.get("status") == "SUCCESS":
            data = j.get("data") or {}
            if tno in (data.get("success") or []):
                return True, batch_no
            return False, f"not in success: {data.get('order_not_found')}"
        return False, j.get("ret_msg") or str(j)
    except Exception as e:
        return False, str(e)


_queue_lock = threading.Lock()
_results_lock = threading.Lock()
_worker_state_lock = threading.Lock()

_queue: deque = deque()
_results: Dict[str, Dict] = {}
_results_order: List[str] = []

_worker_running = False
_worker_thread = None

_blind_cache: Dict[int, Tuple] = {}
_manual_blind: str = ""


def _process_one(tno: str) -> Dict:
    global _blind_cache, _manual_blind

    try:
        detail = get_order_detail(tno)
        s = parse_order(detail)

        order_id = s["order_id"]
        warehouse = s["warehouse"]
        driver_id = s["driver_id"]
        state = s["state"]
        frt = s["failed_reason_type"]
        zipc = s["zip"]
        real_tno = s["tno"] or tno
        storage_info = s.get("storage_info") or ""

        if not order_id or not warehouse:
            return {
                "tno": tno,
                "type": "NA",
                "zip": _norm_zip(zipc),
                "state": state,
                "from_driver": driver_id,
                "to_driver": None,
                "status_updated": False,
                "batch_no": None,
                "message": "系统无记录 (N/A)",
                "warehouse": "",
                "address": "",
                "consignee": "",
                "storage_info": "",
            }

        wh_name = WAREHOUSE_NAME_MAP.get(warehouse, f"Warehouse {warehouse}")

        if state == 203:
            return {
                "tno": tno,
                "type": "DUPLICATE",
                "zip": _norm_zip(zipc),
                "state": state,
                "from_driver": driver_id,
                "to_driver": None,
                "status_updated": False,
                "batch_no": None,
                "message": "状态 203 — 疑似 Duplicate，请人工处理",
                "warehouse": wh_name,
                "address": "",
                "consignee": "",
                "storage_info": "",
            }

        # ── 213 / 230: Stored — transfer to blind batch + assign driver 310991 ──
        if state in (213, 230):
            # get blind batch (manual override or auto-fetch)
            if warehouse not in _blind_cache:
                if _manual_blind:
                    _blind_cache[warehouse] = (_manual_blind, "手动输入")
                else:
                    _blind_cache[warehouse] = get_blind_batch(warehouse)

            batch_no, batch_name = _blind_cache[warehouse]

            msg_parts = [f"storage: {storage_info}" if storage_info else "无storage信息"]

            if not batch_no:
                return {
                    "tno": tno,
                    "type": "STORED",
                    "zip": _norm_zip(zipc),
                    "state": state,
                    "from_driver": driver_id,
                    "to_driver": None,
                    "status_updated": False,
                    "batch_no": None,
                    "message": f"状态{state} 已入库({storage_info}) — 未找到blind批次，请手动处理",
                    "warehouse": wh_name,
                    "address": "",
                    "consignee": s["consignee"],
                    "storage_info": storage_info,
                }

            # transfer to blind batch
            transfer_ok, transfer_msg = do_transfer_to_batch(real_tno, batch_no, warehouse)
            msg_parts.append(
                f"转blind {batch_no}({'成功' if transfer_ok else '失败:' + transfer_msg})"
            )

            # assign to 310991
            assign_ok, assign_msg = False, ""
            if transfer_ok:
                assign_ok, assign_msg = do_assign_driver(order_id, driver_id, STORED_DRIVER)
                msg_parts.append(
                    f"转driver {STORED_DRIVER}({'成功' if assign_ok else '失败:' + assign_msg})"
                )

            return {
                "tno": tno,
                "type": "STORED",
                "zip": _norm_zip(zipc),
                "state": state,
                "from_driver": driver_id,
                "to_driver": STORED_DRIVER if assign_ok else None,
                "status_updated": False,
                "batch_no": batch_no if transfer_ok else None,
                "batch_name": batch_name,
                "message": " | ".join(msg_parts),
                "warehouse": wh_name,
                "address": "",
                "consignee": s["consignee"],
                "storage_info": storage_info,
            }

        is_wrong = (state == 211 and frt == 3)

        if is_wrong:
            if warehouse not in _blind_cache:
                if _manual_blind:
                    _blind_cache[warehouse] = (_manual_blind, "手动输入")
                else:
                    _blind_cache[warehouse] = get_blind_batch(warehouse)

            batch_no, batch_name = _blind_cache[warehouse]

            if not batch_no:
                return {
                    "tno": tno,
                    "type": "WRONG_NO_BATCH",
                    "zip": _norm_zip(zipc),
                    "state": state,
                    "from_driver": driver_id,
                    "to_driver": None,
                    "status_updated": False,
                    "batch_no": None,
                    "message": "Wrong Address — 未找到 blind 批次，请手动输入",
                    "warehouse": wh_name,
                    "address": f"{s['address']} {s['city']} {s['province']} {zipc}".strip(),
                    "consignee": s["consignee"],
                    "storage_info": "",
                }

            transfer_ok, transfer_msg = do_transfer_to_batch(real_tno, batch_no, warehouse)

            driver_998 = get_998_driver(warehouse)
            assign_998_ok = False
            assign_998_msg = ""

            if transfer_ok:
                assign_998_ok, assign_998_msg = do_assign_driver(order_id, driver_id, driver_998)

            msg_parts = [
                f"转 blind {batch_no}（{batch_name}）{'成功' if transfer_ok else '失败:' + transfer_msg}",
            ]
            if transfer_ok:
                msg_parts.append(
                    f"转 driver {driver_998} {'成功' if assign_998_ok else '失败:' + assign_998_msg}"
                )

            return {
                "tno": tno,
                "type": "WRONG_ADDRESS",
                "zip": _norm_zip(zipc),
                "state": state,
                "from_driver": driver_id,
                "to_driver": driver_998 if assign_998_ok else None,
                "status_updated": False,
                "batch_no": batch_no if transfer_ok else None,
                "batch_name": batch_name,
                "message": " | ".join(msg_parts),
                "warehouse": wh_name,
                "address": f"{s['address']} {s['city']} {s['province']} {zipc}".strip(),
                "consignee": s["consignee"],
                "storage_info": "",
            }

        zip5 = _norm_zip(zipc)
        to_driver = route_lookup(zip5)

        if not to_driver:
            return {
                "tno": tno,
                "type": "NO_MATCH",
                "zip": zip5,
                "state": state,
                "from_driver": driver_id,
                "to_driver": None,
                "status_updated": False,
                "batch_no": None,
                "message": f"邮编 {zip5} 未在路线表中找到",
                "warehouse": wh_name,
                "address": "",
                "consignee": "",
                "storage_info": "",
            }

        assign_ok, assign_msg = do_assign_driver(order_id, driver_id, to_driver)

        status_updated = False
        status_msg = ""

        if assign_ok:
            if state not in SKIP_STATUS_UPDATE:
                do_insert_log(order_id, 202)
                upd_ok, upd_msg = do_update_202(order_id, warehouse, real_tno)
                status_updated = upd_ok
                status_msg = f" | 转202 {'成功' if upd_ok else '失败:' + upd_msg}"
            else:
                status_msg = f" | state={state} 无需转202"

        return {
            "tno": tno,
            "type": "OK" if assign_ok else "ERROR",
            "zip": zip5,
            "state": state,
            "from_driver": driver_id,
            "to_driver": to_driver if assign_ok else None,
            "status_updated": status_updated,
            "batch_no": None,
            "message": f"转路线号 {to_driver} {'成功' if assign_ok else '失败:' + assign_msg}{status_msg}",
            "warehouse": wh_name,
            "address": "",
            "consignee": "",
            "storage_info": "",
        }

    except Exception as e:
        return {
            "tno": tno,
            "type": "ERROR",
            "zip": "",
            "state": None,
            "from_driver": None,
            "to_driver": None,
            "status_updated": False,
            "batch_no": None,
            "message": str(e) or "未知错误",
            "warehouse": "",
            "address": "",
            "consignee": "",
            "storage_info": "",
        }


def _worker():
    global _worker_running

    while True:
        tno = None

        with _queue_lock:
            if _queue:
                tno = _queue.popleft()

        if tno is None:
            time.sleep(0.3)
            with _queue_lock:
                if not _queue:
                    with _worker_state_lock:
                        _worker_running = False
                    return
            continue

        with _results_lock:
            if tno in _results:
                _results[tno]["status"] = "processing"

        result = _process_one(tno)
        result["status"] = "done"

        with _results_lock:
            _results[tno] = result

        time.sleep(0.6)


def _ensure_worker():
    global _worker_running, _worker_thread
    with _worker_state_lock:
        if not _worker_running:
            _worker_running = True
            _worker_thread = threading.Thread(target=_worker, daemon=True)
            _worker_thread.start()


app = Flask(__name__, static_folder=".", static_url_path="")


@app.get("/")
def index():
    return send_from_directory(".", "index.html")


@app.get("/api/ping")
def api_ping():
    with _route_lock:
        ok = _route_loaded and bool(_zip_to_route)
        msg = _route_msg
    return jsonify({
        "ok": True,
        "version": APP_VERSION,
        "logged_in": token_mgr.is_ready(),
        "user": CURRENT_USER,
        "route_ok": ok,
        "route_msg": msg,
    })


@app.post("/api/login")
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    u = (data.get("username") or "").strip()
    p = data.get("password") or ""

    if not u or not p:
        return jsonify({"ok": False, "msg": "请输入账号密码"})

    try:
        token_mgr.set_credentials(u, p)
        token_mgr.get_token()
        ok, msg, cnt = load_routes()
        return jsonify({
            "ok": True,
            "user": CURRENT_USER,
            "route_ok": ok,
            "route_msg": msg,
            "route_cnt": cnt,
        })
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)})


@app.post("/api/reload_routes")
def api_reload_routes():
    ok, msg, cnt = load_routes()
    return jsonify({"ok": ok, "msg": msg, "count": cnt})


@app.post("/api/set_blind")
def api_set_blind():
    global _manual_blind, _blind_cache
    data = request.get_json(force=True, silent=True) or {}
    _manual_blind = (data.get("blind_batch") or "").strip()
    _blind_cache = {}
    return jsonify({"ok": True, "blind_batch": _manual_blind})


@app.post("/api/enqueue")
def api_enqueue():
    if not token_mgr.is_ready():
        return jsonify({"ok": False, "msg": "未登录"})

    with _route_lock:
        if not (_route_loaded and _zip_to_route):
            return jsonify({"ok": False, "msg": f"路线表未加载: {_route_msg}"})

    data = request.get_json(force=True, silent=True) or {}
    raw = (data.get("tnos") or "").replace(",", "\n")
    tnos = [x.strip() for x in raw.split() if x.strip()]

    added = []

    with _queue_lock, _results_lock:
        for tno in tnos:
            if tno not in _results:
                _results[tno] = {
                    "tno": tno,
                    "type": "PENDING",
                    "status": "pending",
                    "message": "等待中…",
                    "zip": "",
                    "state": None,
                    "from_driver": None,
                    "to_driver": None,
                    "status_updated": False,
                    "batch_no": None,
                    "warehouse": "",
                    "address": "",
                    "consignee": "",
                    "storage_info": "",
                }
                _results_order.append(tno)
                _queue.append(tno)
                added.append(tno)

    if added:
        _ensure_worker()

    return jsonify({"ok": True, "added": added, "queued": len(added)})


@app.get("/api/status")
def api_status():
    with _results_lock:
        results = [_results[t] for t in _results_order if t in _results]
    with _queue_lock:
        q_len = len(_queue)

    total = len(results)
    pending = sum(1 for r in results if r["status"] == "pending")
    processing = sum(1 for r in results if r["status"] == "processing")
    done = sum(1 for r in results if r["status"] == "done")
    ok_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] == "OK")
    wrong_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] in ("WRONG_ADDRESS", "WRONG_NO_BATCH"))
    err_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] in ("ERROR", "NO_MATCH"))
    na_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] == "NA")
    dup_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] == "DUPLICATE")
    stored_cnt = sum(1 for r in results if r["status"] == "done" and r["type"] == "STORED")

    return jsonify({
        "ok": True,
        "queue_len": q_len,
        "total": total,
        "pending": pending,
        "processing": processing,
        "done": done,
        "ok_cnt": ok_cnt,
        "wrong_cnt": wrong_cnt,
        "err_cnt": err_cnt,
        "na_cnt": na_cnt,
        "dup_cnt": dup_cnt,
        "stored_cnt": stored_cnt,
        "results": list(reversed(results)),
    })


@app.post("/api/clear")
def api_clear():
    global _blind_cache, _manual_blind
    with _queue_lock:
        _queue.clear()
    with _results_lock:
        _results.clear()
        _results_order.clear()
    _blind_cache = {}
    _manual_blind = ""
    return jsonify({"ok": True})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "9000"))
    load_routes()
    app.run(host="0.0.0.0", port=port, debug=False)
